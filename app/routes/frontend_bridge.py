from fastapi import APIRouter, HTTPException, UploadFile, File
from typing import Dict, Any, List

from app.services.pre_advising_service import PreAdvisingService
from app.services.logout import LogoutService
from app.services.auth import AuthService
from app.db import (
    students_col,
    users_col,
    sections_col,
    schedule_col,
    curriculum_col,
    academic_records_col,
)

router = APIRouter(prefix="/bridge", tags=["frontend-bridge"])

logout_service = LogoutService()
auth_service = AuthService()
current_user_username: str = None


@router.post("/login")
async def bridge_login(credentials: Dict[str, str]) -> Dict[str, Any]:
    try:
        username = credentials.get("username")
        password = credentials.get("password")
        if not username or not password:
            raise HTTPException(status_code=400, detail="Username and password required")
        user = users_col.find_one({"username": username})
        if not user:
            return {"success": False, "error": "Invalid credentials"}
        if user.get("password") != password:
            return {"success": False, "error": "Invalid credentials"}
        user.pop("password", None)
        user["_id"] = str(user.get("_id"))
        user_response = {
            "username":  user.get("username") or "",
            "full_name": user.get("full_name") or user.get("fullName") or user.get("name") or "",
            "gsuite":    user.get("gsuite") or user.get("gSuite") or user.get("email") or "",
            "position":  user.get("position") or user.get("role") or user.get("Position") or "",
        }
        try:
            global current_user_username
            current_user_username = user.get("username")
        except Exception:
            pass
        return {"success": True, "user": user_response}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/logout")
async def bridge_logout() -> Dict[str, str]:
    try:
        auth_service.handle_logout()
        return {"success": True, "message": "Logged out successfully"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/dashboard")
async def bridge_dashboard() -> Dict[str, Any]:
    try:
        cursor = students_col.find({})
        students = list(cursor)
        result = []
        for s in students:
            s["_id"] = str(s.get("_id"))
            result.append({
                **s,
                "name":   s.get("name")   or s.get("full_name") or "",
                "email":  s.get("email")  or s.get("gsuite")    or s.get("email_address") or "",
                "number": s.get("number") or s.get("student_no") or s.get("student_number") or "",
            })

        total           = len(result)
        total_regular   = sum(1 for s in result if (s.get("status") or "").lower() == "regular")
        total_irregular = sum(1 for s in result if (s.get("status") or "").lower() == "irregular")

        year_counts = {"1": 0, "2": 0, "3": 0, "4": 0}
        for s in result:
            year = str(
                s.get("year_level") or s.get("year") or
                s.get("yearLevel")  or s.get("Year Level") or ""
            ).strip()
            if year in year_counts:
                year_counts[year] += 1

        section_counts = {"1": 0, "2": 0, "3": 0, "4": 0}
        for sec in sections_col.find({}):
            year = str(
                sec.get("year_level") or sec.get("year") or
                sec.get("yearLevel")  or sec.get("Year Level") or ""
            ).strip()
            if year in section_counts:
                section_counts[year] += 1

        return {"success": True, "data": {
            "students":        result,
            "total_students":  total,
            "total_regular":   total_regular,
            "total_irregular": total_irregular,
            "year_counts":     year_counts,
            "section_counts":  section_counts,
        }}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/year-levels")
async def bridge_year_levels() -> Dict[str, Any]:
    try:
        def count_sections(level):
            return sections_col.count_documents({
                "$or": [
                    {"year_level": level},
                    {"year_level": str(level)},
                    {"year": level},
                    {"year": str(level)},
                ]
            })

        year_levels = [
            {"year_level": 1, "name": "BSIT 1st Year", "section_count": count_sections(1)},
            {"year_level": 2, "name": "BSIT 2nd Year", "section_count": count_sections(2)},
            {"year_level": 3, "name": "BSIT 3rd Year", "section_count": count_sections(3)},
            {"year_level": 4, "name": "BSIT 4th Year", "section_count": count_sections(4)},
            {
                "year_level": "irregular",
                "name": "Irregular Students",
                "student_count": students_col.count_documents({
                    "status": {"$regex": "irregular", "$options": "i"}
                }),
            },
        ]
        return {"success": True, "data": year_levels}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/sections/{year_level}")
async def bridge_sections(year_level) -> Dict[str, Any]:
    try:
    
        if str(year_level).lower() == "irregular":
            return {
                "success": True,
                "data": [],
                "message": "Irregular students do not have sections. Use /students/irregular to get irregular student list."
            }
        
        
        try:
            year_level_int = int(year_level)
        except (ValueError, TypeError):
            return {"success": False, "error": f"Invalid year level: {year_level}"}
        
        
        cursor = sections_col.find({
            "$or": [
                {"year_level": year_level_int},
                {"year_level": str(year_level_int)},
                {"year":       year_level_int},
                {"year":       str(year_level_int)},
            ]
        })
        sections = list(cursor)

        result = []
        for s in sections:
            s["_id"] = str(s.get("_id"))
            result.append({
                "name":       s.get("name") or s.get("section_name") or s.get("section") or "",
                "year_level": s.get("year_level") or s.get("year") or year_level_int,
                "active":     s.get("active") or False,
            })

        return {"success": True, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/students/irregular")
async def bridge_irregular_students() -> Dict[str, Any]:
    try:
        cursor = students_col.find({"status": {"$regex": "irregular", "$options": "i"}})
        students = list(cursor)
        result = []
        for s in students:
            s["_id"] = str(s.get("_id"))
            result.append({
                **s,
                "name":   s.get("name")   or s.get("full_name") or "",
                "email":  s.get("email")  or s.get("gsuite")    or s.get("email_address") or "",
                "number": s.get("number") or s.get("student_no") or s.get("student_number") or "",
            })
        return {"success": True, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/students/by-category")
async def bridge_students_by_category(category: str = None) -> Dict[str, Any]:
    
    try:
        if not category:
            return {"success": False, "error": "category parameter is required"}
        
        if category.lower() == "irregular":
            
            cursor = students_col.find({"status": {"$regex": "irregular", "$options": "i"}})
            students = list(cursor)
            result = []
            for s in students:
                s["_id"] = str(s.get("_id"))
                result.append({
                    **s,
                    "name":   s.get("name")   or s.get("full_name") or "",
                    "email":  s.get("email")  or s.get("gsuite")    or s.get("email_address") or "",
                    "number": s.get("number") or s.get("student_no") or s.get("student_number") or "",
                })
            return {"success": True, "data": result, "category": "irregular"}
        else:
            return {"success": False, "error": f"Unknown category: {category}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/students")
async def bridge_all_students() -> Dict[str, Any]:
    try:
        cursor = students_col.find({})
        students = list(cursor)
        result = []
        for s in students:
            s["_id"] = str(s.get("_id"))
            result.append({
                **s,
                "name":   s.get("name")   or s.get("full_name") or "",
                "email":  s.get("email")  or s.get("gsuite")    or s.get("email_address") or "",
                "number": s.get("number") or s.get("student_no") or s.get("student_number") or "",
            })
        return {"success": True, "data": {"students": result}}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/students/{section}")
async def bridge_students_by_section(section: str) -> Dict[str, Any]:
    try:
        
        cursor = students_col.find({
            "$and": [
                {
                    "$or": [
                        {"section": {"$regex": f"^{section}$", "$options": "i"}},
                        {"Section": {"$regex": f"^{section}$", "$options": "i"}},
                    ]
                },
                {
                    "$or": [
                        {"status": {"$regex": "^regular$", "$options": "i"}},
                        {"status": {"$exists": False}},  # Treat missing status as regular
                    ]
                }
            ]
        })
        students = list(cursor)
        result = []
        for s in students:
            s["_id"] = str(s.get("_id"))
            result.append({
                **s,
                "name":   s.get("name")   or s.get("full_name") or "",
                "email":  s.get("email")  or s.get("gsuite")    or s.get("email_address") or "",
                "number": s.get("number") or s.get("student_no") or s.get("student_number") or "",
            })

        return {"success": True, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/student-search")
async def bridge_student_search(student_no: str = "") -> Dict[str, Any]:
    try:
        query = {"$or": [
            {"student_no": {"$regex": student_no, "$options": "i"}},
            {"number": {"$regex": student_no, "$options": "i"}},
            {"student_number": {"$regex": student_no, "$options": "i"}},
        ]}
        cursor = students_col.find(query)
        students = list(cursor)
        result = []
        for s in students:
            s["_id"] = str(s.get("_id"))
            result.append({
                **s,
                "name":   s.get("name")   or s.get("full_name") or "",
                "email":  s.get("email")  or s.get("gsuite")    or s.get("email_address") or "",
                "number": s.get("number") or s.get("student_no") or s.get("student_number") or "",
            })
        return {"success": True, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/student/{student_no}")
async def bridge_student_details(student_no: str) -> Dict[str, Any]:
    try:
        student = students_col.find_one({
            "$or": [
                {"student_no": student_no},
                {"number":     student_no},
                {"student_number": student_no},
            ]
        })
        if student:
            student["_id"]    = str(student.get("_id"))
            student["number"] = student.get("number") or student.get("student_no") or ""
            student["name"]   = student.get("name")   or student.get("full_name")  or ""
            student["email"]  = student.get("email")  or student.get("gsuite")     or ""

        records_doc = academic_records_col.find_one({
            "$or": [
                {"student_no":     student_no},
                {"number":         student_no},
                {"student_number": student_no},
            ]
        })
        raw_grades = []
        if records_doc:
            raw_grades = records_doc.get("records") or records_doc.get("grades") or []
            if not raw_grades and records_doc.get("subject_code"):
                cursor = academic_records_col.find({
                    "$or": [
                        {"student_no":     student_no},
                        {"number":         student_no},
                        {"student_number": student_no},
                    ]
                })
                raw_grades = list(cursor)

        curriculum_index = {}
        try:
            curr_docs = list(curriculum_col.find({}))
            for doc in curr_docs:
                for year in doc.get("years", []):
                    semesters = year.get("semesters") or {}
                    for sem_key, sem_subjects in semesters.items():
                        if isinstance(sem_subjects, list):
                            for subj in sem_subjects:
                                code = (
                                    subj.get("subject_code") or subj.get("code") or
                                    subj.get("subjectCode") or subj.get("course_code") or ""
                                )
                                if code:
                                    normalized_code = str(code).strip().upper()
                                    curriculum_index[normalized_code] = {
                                        "title": (
                                            subj.get("title") or subj.get("name") or
                                            subj.get("subject_title") or subj.get("descriptive_title") or
                                            subj.get("description") or ""
                                        ),
                                        "units": (
                                            subj.get("units") or subj.get("unit") or
                                            subj.get("credit_units") or subj.get("Units") or ""
                                        ),
                                    }
        except Exception:
            pass

        def normalized_status(record: Dict[str, Any]) -> str:
            status_raw = str(record.get("status") or "").strip().lower()
            credited_flag = record.get("credited") is True or status_raw == "credited"
            if credited_flag:
                return "credited"
            if status_raw in {"completed", "passed"}:
                return "completed"

            final_grade_value = record.get("grade") or record.get("final_grade") or record.get("finalGrade") or record.get("final")
            if isinstance(final_grade_value, str):
                grade_str = final_grade_value.strip().upper()
                if grade_str in {"INC", "DRP", ""}:
                    return "missing"
                try:
                    final_grade_value = float(grade_str)
                except ValueError:
                    return "missing"
            if isinstance(final_grade_value, (int, float)):
                try:
                    final_grade_value = float(final_grade_value)
                except Exception:
                    return "missing"
                if 1.0 <= final_grade_value <= 3.0:
                    return "completed"
                if final_grade_value == 5.0 or final_grade_value == 0.0:
                    return "missing"
            return "missing"

        normalized_grades = []
        for g in raw_grades:
            code = g.get("subject_code") or g.get("code") or g.get("subjectCode") or ""
            normalized_code = str(code).strip().upper()
            curr_info = curriculum_index.get(normalized_code, {})
            final_grade_value = g.get("grade") or g.get("final_grade") or g.get("finalGrade") or g.get("final") or ""
            normalized_grades.append({
                "code":       code,
                "title":      curr_info.get("title") or g.get("title") or "",
                "units":      curr_info.get("units") or g.get("units") or g.get("unit") or g.get("credit_units") or "",
                "finalGrade": str(final_grade_value),
                "prelim":     str(g.get("prelim")   or g.get("prelim_grade")    or ""),
                "midterm":    str(g.get("midterm")  or g.get("midterm_grade")   or ""),
                "preFinal":   str(g.get("preFinal") or g.get("pre_final")       or ""),
                "final":      str(g.get("final")    or g.get("final_period")    or ""),
                "instructor": g.get("instructor") or g.get("teacher") or "",
                "semester":   str(g.get("semesters") or g.get("semester") or ""),
                "schoolYear": str(g.get("school_year") or ""),
                "status":     normalized_status(g),
            })

        return {"success": True, "data": {"student": student, "grades": normalized_grades}}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/pre-advising")
async def bridge_pre_advising(payload: Dict[str, Any]) -> Dict[str, Any]:
    try:
        pre_advising = PreAdvisingService(None)
        student_no = payload.get("student_no")
        year_level = payload.get("year_level")
        semester   = payload.get("semester")
        status     = payload.get("status")
        recommendations = await pre_advising.generate_pre_advising(
            student_no, year_level, semester, status
        )
        return {"success": True, "data": recommendations}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/schedule/all")
async def bridge_schedule_all() -> Dict[str, Any]:
    try:
        cursor   = schedule_col.find({})
        schedule = list(cursor)
        for s in schedule:
            s["_id"] = str(s.get("_id"))
        return {"success": True, "data": schedule}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/schedule/{section}")
async def bridge_schedule_section(section: str, semester: str = "") -> Dict[str, Any]:
    try:
        section_pattern = section.replace(" ", "\\s*")
        section_query   = {"$regex": section_pattern, "$options": "i"}
        query = {
            "$or": [
                {"section":      section_query},
                {"Section":      section_query},
                {"section_name": section_query},
            ]
        }
        if semester:
            semester_query = {"$regex": semester, "$options": "i"}
            query["$and"] = [{"$or": [
                {"semester":    semester_query},
                {"Semester":    semester_query},
                {"school_year": semester_query},
            ]}]

        docs = list(schedule_col.find(query))

        
        if not docs and semester:
            base_query = {
                "$or": [
                    {"section":      section_query},
                    {"Section":      section_query},
                    {"section_name": section_query},
                ]
            }
            docs = list(schedule_col.find(base_query))

        result = []
        for doc in docs:
            doc["_id"] = str(doc.get("_id"))
            time_value = (
                doc.get("time") or doc.get("Time") or doc.get("time_start") or doc.get("start_time") or ""
            )
            if not time_value:
                start = doc.get("start_time") or doc.get("begin_time") or ""
                end = doc.get("end_time") or doc.get("finish_time") or ""
                if start and end:
                    time_value = f"{start} - {end}"
            result.append({
                "subjectCode": doc.get("subject_code") or doc.get("subjectCode") or doc.get("code") or doc.get("description") or "",
                "unit":        doc.get("units")  or doc.get("unit")  or doc.get("Units")       or "",
                "hours":       doc.get("hours")  or doc.get("Hours")                            or "",
                "time":        time_value,
                "days":        doc.get("days")   or doc.get("Days")  or doc.get("day")          or "",
                "room":        doc.get("room")   or doc.get("Room")  or doc.get("room_no")      or "",
                "section":     doc.get("section") or doc.get("Section")                         or "",
                "instructor":  doc.get("instructor") or doc.get("teacher") or doc.get("faculty") or "",
                "semester":    doc.get("semester")   or doc.get("Semester")                     or "",
                "schoolYear":  doc.get("school_year") or doc.get("schoolYear")                  or "",
            })

        return {"success": True, "data": result}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/curriculum")
async def bridge_curriculum() -> Dict[str, Any]:
    from app.database import curriculum_collection
    cursor = curriculum_collection.find({})
    docs = await cursor.to_list(length=None)
    year_names = {1: "First Year", 2: "Second Year", 3: "Third Year", 4: "Fourth Year"}
    result: Dict[str, Any] = {}
    for doc in docs:
        for year in doc.get("years", []):
            year_label = year_names.get(year.get("year_level"), f"Year {year.get('year_level')}")
            result[year_label] = {"First Semester": [], "Second Semester": []}
            semesters = year.get("semesters", {})
            for s in semesters.get("1", []):
                result[year_label]["First Semester"].append({
                    "code": s.get("subject_code"), "title": s.get("title"),
                    "lec":  s.get("lec"),          "lab":   s.get("lab"),
                    "units": s.get("units"),
                })
            for s in semesters.get("2", []):
                result[year_label]["Second Semester"].append({
                    "code": s.get("subject_code"), "title": s.get("title"),
                    "lec":  s.get("lec"),          "lab":   s.get("lab"),
                    "units": s.get("units"),
                })
    return {"success": True, "data": result}


@router.get("/profile")
async def bridge_profile() -> Dict[str, Any]:
    try:
        if not current_user_username:
            raise HTTPException(status_code=401, detail="Not logged in")
        profile = users_col.find_one({"username": current_user_username})
        if not profile:
            raise HTTPException(status_code=404, detail="Profile not found")
        profile.pop("password", None)
        profile["_id"] = str(profile.get("_id"))
        return {"success": True, "data": profile}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.put("/profile")
async def bridge_update_profile(updates: Dict[str, str]) -> Dict[str, Any]:
    try:
        if not current_user_username:
            raise HTTPException(status_code=401, detail="Not logged in")
        allowed  = {"full_name", "username", "gsuite", "position"}
        filtered = {k: v for k, v in updates.items() if k in allowed}
        users_col.update_one({"username": current_user_username}, {"$set": filtered})
        updated = users_col.find_one({"username": filtered.get("username", current_user_username)})
        if updated:
            updated.pop("password", None)
            updated["_id"] = str(updated.get("_id"))
        return {"success": True, "data": updated}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/profile/upload")
async def bridge_upload_profile_pic(file: UploadFile = File(...)) -> Dict[str, Any]:
    try:
        import tempfile, os
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name
        users_col.update_one(
            {"username": current_user_username or "admin"},
            {"$set": {"profile_picture": tmp_path}}
        )
        updated = users_col.find_one({"username": current_user_username or "admin"})
        if updated:
            updated.pop("password", None)
            updated["_id"] = str(updated.get("_id"))
        return {"success": True, "data": updated}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.delete("/profile/picture")
async def bridge_delete_profile_pic() -> Dict[str, Any]:
    try:
        users_col.update_one(
            {"username": current_user_username or "admin"},
            {"$unset": {"profile_picture": ""}}
        )
        updated = users_col.find_one({"username": current_user_username or "admin"})
        if updated:
            updated.pop("password", None)
            updated["_id"] = str(updated.get("_id"))
        return {"success": True, "data": updated}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/logout/confirm")
async def bridge_logout_confirm() -> Dict[str, bool]:
    return {"show_confirmation": logout_service.is_showing_confirmation()}


@router.post("/logout/yes")
async def bridge_logout_confirm_yes() -> Dict[str, str]:
    try:
        logout_service.handle_confirm_logout(lambda: auth_service.handle_logout())
        return {"success": True, "message": "Logged out"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/logout/no")
async def bridge_logout_cancel() -> Dict[str, str]:
    try:
        logout_service.handle_cancel_logout()
        return {"success": True, "message": "Logout cancelled"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/logout/click")
async def bridge_logout_click() -> Dict[str, str]:
    try:
        logout_service.handle_logout_click()
        return {"success": True, "message": "Logout confirmation shown"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/test-user")
async def test_user():
    user = users_col.find_one({"username": "admin001"})
    if user:
        user["_id"] = str(user["_id"])
        return {"found": True, "data": dict(user)}
    return {"found": False}


@router.post("/upload-grades")
async def bridge_upload_grades(file: UploadFile = File(...)) -> Dict[str, Any]:
    try:
        import io
        import csv

        filename = file.filename or ""
        content  = await file.read()

        rows = []

        if filename.endswith(".csv"):
            decoded = content.decode("utf-8-sig")
            reader  = csv.DictReader(io.StringIO(decoded))
            rows    = [row for row in reader]

        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            try:
                import openpyxl
                wb   = openpyxl.load_workbook(io.BytesIO(content))
                ws   = wb.active
                headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            except ImportError:
                return {"success": False, "error": "openpyxl is not installed. Run: pip install openpyxl"}

        else:
            return {"success": False, "error": "Unsupported file type. Please upload a .csv or .xlsx file."}

        if not rows:
            return {"success": False, "error": "File is empty or has no data rows."}

        updated = 0
        inserted = 0
        errors = []

        for i, row in enumerate(rows):
            student_no = (
                row.get("student_no") or row.get("student_number") or
                row.get("Student No") or row.get("Student Number") or
                row.get("ID") or ""
            ).strip()

            subject_code = (
                row.get("subject_code") or row.get("code") or
                row.get("Subject Code") or row.get("Code") or ""
            ).strip()

            if not student_no or not subject_code:
                errors.append(f"Row {i+2}: missing student_no or subject_code")
                continue

            grade_record = {
                "subject_code": subject_code,
                "grade":        row.get("grade") or row.get("final_grade") or row.get("Grade") or "",
                "status":       row.get("status") or row.get("Status") or "completed",
                "school_year":  row.get("school_year") or row.get("School Year") or "",
                "semesters":    row.get("semesters") or row.get("semester") or row.get("Semester") or "",
                "prelim":       row.get("prelim") or row.get("Prelim") or "",
                "midterm":      row.get("midterm") or row.get("Midterm") or "",
                "preFinal":     row.get("preFinal") or row.get("pre_final") or row.get("Pre-Final") or "",
                "final":        row.get("final") or row.get("Final") or "",
            }

            existing = academic_records_col.find_one({
                "$or": [
                    {"student_no": student_no},
                    {"number": student_no},
                    {"student_number": student_no},
                ]
            })

            if existing:
                records = existing.get("records") or []
                
                found = False
                for rec in records:
                    if rec.get("subject_code", "").upper() == subject_code.upper():
                        rec.update(grade_record)
                        found = True
                        break
                if not found:
                    records.append(grade_record)

                academic_records_col.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {"records": records}}
                )
                updated += 1
            else:
                academic_records_col.insert_one({
                    "student_no": student_no,
                    "records": [grade_record]
                })
                inserted += 1

        return {
            "success": True,
            "message": f"Upload complete. {updated} student(s) updated, {inserted} new record(s) inserted.",
            "updated": updated,
            "inserted": inserted,
            "errors": errors,
        }

    except Exception as e:
        return {"success": False, "error": str(e)}